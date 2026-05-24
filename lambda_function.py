import boto3
import json
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
import time # Timer
from datetime import datetime, timezone
from botocore.exceptions import ClientError
from croniter import croniter
# If you need more information about configurations
# or implementing the sample code, visit the AWS docs:
# https://aws.amazon.com/developer/language/python/


def results_from_file(bucket_name, region_name):
    s3 = boto3.resource(service_name='s3', region_name=region_name)
    bucket = s3.Bucket(bucket_name)
    
    local_path = "/tmp/report.json"
    
    try:
        bucket.download_file("report.json", local_path)
        print_and_log("report.json downloaded from S3", level="INFO")
    except ClientError as e:
        print_and_log(f"Failed to download report.json: {e}", level="ERROR")
        return None

    with open(local_path, "r") as f:
        data = json.load(f)
    
    print_and_log("Data loaded from report.json", level="INFO")
    return data

def next_execution_time(schedule_name: str, group_name: str = None):
    client = boto3.client('scheduler')
    params = {'Name': schedule_name}
    if group_name:
        params['GroupName'] = group_name

    try:
        resp = client.get_schedule(**params)
        raw = resp['ScheduleExpression']  # e.g. "cron(0 8 ? * MON *)"

        if raw.lower().startswith('cron(') and raw.endswith(')'):
            expr = raw[5:-1]
        else:
            expr = raw

        expr = expr.replace('?', '*').strip()

        now = datetime.now(timezone.utc)
        it  = croniter(expr, now)
        next_dt = it.get_next(datetime)
        print_and_log(f"Next schedule will be in {next_dt.strftime('%Y-%m-%d %H:%M:%S')} (UTC)", level="INFO")
        return next_dt

    except Exception as e:
        print_and_log(f"Error getting next schedule time: {e}", level="ERROR")
        return None


def sns_undiligent(topic_arn, timestamp, new_undil, all_undil, fixed_undil):
    sns = boto3.client('sns')

    subject = f"[Wizer] Report about undiligent users"
    body_lines = [
        f"Report time (UTC): {timestamp}",
        "",
        f"Users who are not undiligent anymore ({len(fixed_undil)}): ",
        *(f"  •  {i}" for i in fixed_undil or ["None"]),
        "",
        f"New undiligent users ({len(new_undil)}):",
        *(f"  •  {u}" for u in new_undil or ["None"]),
        "",
        f"All undiligent users now ({len(all_undil)}):",
        *(f"  •  {i}" for i in all_undil or ["None"]),
    ]
    message = "\n".join(body_lines)

    sns.publish(
        TopicArn=topic_arn,
        Subject=subject,
        Message=message
    )
    print_and_log(f"SNS notification sent to {topic_arn}", "INFO")


def get_secret(secret_name, region_name):

    # Create a Secrets Manager client
    session = boto3.session.Session()
    client = session.client(
        service_name='secretsmanager',
        region_name=region_name
    )

    try:
        get_secret_value_response = client.get_secret_value(
            SecretId=secret_name
        )
    except ClientError as e:
        # For a list of exceptions thrown, see
        # https://docs.aws.amazon.com/secretsmanager/latest/apireference/API_GetSecretValue.html
        raise e

    secret = get_secret_value_response['SecretString']
    return secret

def print_and_log(message, level="UNKNOWN"):

    error_types = ["DEBUG", "TRACE", "INFO", "WARNING", "ERROR", "CRITICAL"]
    
    if level not in error_types:
        level = "UNKNOWN"

    print(f"| {level} | {message}")

def excel(diligent_users, undiligent_users, mixed_users, users_dict, excel_filename, excel_filename_full, bucket_name, region_name):

    new_undil_list = []
    not_more_undil = []

    s3 = boto3.resource(
        service_name='s3',
        region_name=region_name
        )
    bucket=s3.Bucket(bucket_name)
    try:
        bucket.Object(excel_filename).load()
        bucket.download_file(excel_filename, excel_filename_full)
        print_and_log("Excel file exists in s3", level="INFO")
        wb = load_workbook(excel_filename_full)
    except ClientError as e:
        error_code = e.response.get("Error", {}).get("Code", "")
        if error_code in ("404"):
            wb = Workbook()
        else:
            raise e

    sheet_name = datetime.now(timezone.utc).strftime("%Y-%m-%d %H-%M-%S (UTC)")
    sheet = wb.create_sheet(title=sheet_name)

    wb._sheets = sorted(wb._sheets, key=lambda ws: ws.title, reverse=True)
   
    #Get previous sheet name (if we have more than one)
    if len(wb.sheetnames) > 1:
        prev_sheet_name = wb.sheetnames[1] #-1
        print_and_log(f"Previous sheet name is: {prev_sheet_name}", level="DEBUG")
    else:
        prev_sheet_name = None

    sheet["A1"] = "Diligent Users"
    sheet["B1"] = "Undiligent Users"
    sheet["C1"] = "Other Users"
    sheet["D1"] = "Comparison"
    sheet["E1"] = "Undiligent fixed compared to Previous Sheet"

    max_rows = max(len(diligent_users), len(undiligent_users), len(mixed_users))
    for i in range(max_rows):
        row_index = i+2
        if i < len(diligent_users):
            sheet.cell(row=i+2, column=1, value=diligent_users[i])
        if i < len(undiligent_users):
            sheet.cell(row=i+2, column=2, value=undiligent_users[i])
        if i < len(mixed_users):

            user_courses = users_dict[mixed_users[i]]['all_user_courses'] # Course list of specific user
            # Counts completed courses...
            completed_courses = 0
            for course in user_courses:
                if course['progress'] == 100:
                    completed_courses += 1
            remaining_courses = len(user_courses) - completed_courses #Remaining courses = Total - completed
            courses_info = f"{mixed_users[i]} (Courses completed: {completed_courses}, Courses left: {remaining_courses})"
            sheet.cell(row=i+2, column=3, value=courses_info)

    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = max(max_length + 2, 20)
        sheet.column_dimensions[col_letter].width = adjusted_width


    results_sheets = [name for name in wb.sheetnames]
    if len (results_sheets) > 5:
        oldest = results_sheets[-1]
        wb.remove(wb[oldest])
        print_and_log(f"Oldest sheet was removed: {oldest}", level="INFO")

    if prev_sheet_name:
        prev_sheet = wb[prev_sheet_name]

        prev_undil = []
        for row in prev_sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
            for cell in row:
                if cell:
                    prev_undil.append(str(cell))

        new_undil = []
        for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
            for cell in row:
                if cell:
                    new_undil.append(str(cell))

        not_more_undil = [user for user in prev_undil if user not in new_undil]
        print_and_log(f"Not more undil list length: {len(not_more_undil)}", level="INFO")

        new_undil_list = [user for user in new_undil if user not in prev_undil]
        print_and_log(f"New undil list length: {len(new_undil_list)}", level="INFO")

        sheet["E2"] = f"Fixed: {len(not_more_undil)}; New: {len(new_undil_list)}"
        sheet["E3"] = "These users are not undiligent anymore: " + (", ".join(not_more_undil) if not_more_undil else "None")
        sheet["E4"] = "These users are new undiligent: " + (", ".join(new_undil_list) if new_undil_list else "None")

    try:
        wb.save(excel_filename_full)
        print_and_log(f"File saved succesfully into {excel_filename_full} !", level="INFO")

        bucket.upload_file(excel_filename_full, excel_filename)
        print_and_log(f"File was saved into s3 bucket: {excel_filename}", level="INFO")

    except PermissionError:
        print_and_log("ERROR: File is busy! Please close the file and try again.", level="ERROR")

    except Exception as e2:
        print_and_log("A problem occured with saving file!", level="ERROR")
        raise e2
    
    return new_undil_list, undiligent_users, not_more_undil

def user_analysis(user, user_data):
            
    user_verdict = ""
    progress_list = [course.get('progress', 0) for course in user_data.get("all_user_courses", [])] #Retrieve courses from dictionary
    course_status_list = [course.get('status', "Unknown") for course in user_data.get("all_user_courses", [])]

    if min(progress_list) == 100:
        user_verdict = "diligent"
    elif (max(progress_list) == 0 or all(status != "Completed" for status in course_status_list )):
        user_verdict = "undiligent"
    else:
        user_verdict = "mixed"

    return user_verdict, user


# status code 500, 503, 504, 429, 404
def wizer_api(url, secret_name, region_name, max_retries=3, delay=5):
    api_key = get_secret(secret_name, region_name)
    valid_codes = [404, 429, 500, 503, 504]
    attempt = 1
    while attempt <= max_retries:
        try:
            #Sends api key
            wizer_report_response = requests.get(url, headers={
                "apiKey": api_key
                })
            if wizer_report_response.status_code == 200:
                print_and_log(f"API request succeeded on attempt {attempt}", level="INFO")
                return wizer_report_response
            if wizer_report_response.status_code in valid_codes:
                print_and_log(f"API responded with status {wizer_report_response.status_code} on attempt {attempt}", level="WARNING")
            else:
                print_and_log(f"API responded with status {wizer_report_response.status_code} on attempt {attempt}", level="WARNING")
                print_and_log("Response code is not retriable! Giving up.", level="CRITICAL")
                return None
        except Exception as e:
            print_and_log(f"API call failed on attempt {attempt}: {str(e)}", level="ERROR")
            return None

        attempt += 1
        
        if attempt <= max_retries:
            print_and_log(f"Retrying in {delay} seconds...", level="INFO")
            time.sleep(delay)
            delay *= 1.5
        else:
            print_and_log("Max retry attempts reached. Giving up.", level="CRITICAL")
            print_and_log("Wizer API failed after multiple retries.", level="CRITICAL")
    return None

def lambda_handler(event, context):
    
    print_and_log("Lambda started...", level="INFO")
    
    output_directory = "/tmp/"
    log_filename = "/tmp/py_log_manual.log"
    url = "https://api.wizer-training.com/api/v1/external/reports/master_report"
    region_name = "eu-central-1"
    bucket_name = "bakalaurabucket"

    secret_name = "bakalaurasecret"
    topic_arn = "arn:aws:sns:eu-central-1:989864147516:bakalauratopic"

    excel_filename = "results.xlsx"
    excel_filename_full = os.path.join(output_directory,excel_filename)

    from_file = True
    print_and_log("Main processing started...", level="INFO")
    
    start_time = time.time()

    if from_file:
        print_and_log("Loading data from file (from_file=True)", level="INFO")
        wizer_report_data = results_from_file(bucket_name, region_name)
        if wizer_report_data is None:
            return {
                'statusCode': 500,
                'body': 'Failed to load data from report.json'
            }
    else:
        try:
            secret = get_secret(secret_name, region_name)
            resp = wizer_api(url, secret_name, region_name)
            
            if resp is None:
                print_and_log("No response from API after retries.", level="ERROR")
                return {
                    'statusCode': 503,
                    'body': 'Function execution stopped due to no response from Wizer API'
                }
            elif hasattr(resp, "status_code") and resp.status_code == 200:
                wizer_report_data = resp.json()
                print_and_log("Status code is 200", level="INFO")
            else:
                print_and_log(f"Unexpected API response: {resp.status_code}", level="ERROR")
                return {
                    'statusCode': 503,
                    'body': 'Wizer API did not return a successful response'
                }
        except Exception as e:
            print_and_log(f"API error: {e}", level="ERROR")
            return {
                    'statusCode': 503,
                    'body': 'Function execution stopped due to no response from Wizer API'
                }


    wizer_report_api_response = {}

    if len(wizer_report_api_response) > 0:
        print(f"Status Code: {wizer_report_api_response.text}", level="DEBUG")
        print(f"Status Code: {wizer_report_api_response.status_code}", level="DEBUG")

 
    diligent_users = []
    undiligent_users = []
    mixed_users = []
    all_users_cnt = 0

    #if wizer_report_response:
    if len(wizer_report_data) > 0:
        #print_and_log("Status code is 200!", level="INFO")
        user_progress_list = wizer_report_data.get("userProgress", [])

        users_dict = {}



        for record in user_progress_list:
            email = record.get("email", "dummy_email@1nce.com")
            first_name = record.get("firstName", "")
            last_name = record.get("lastName", "")
            status = record.get("status","Unknown")
            progress = record.get("progress", 0)
            course = record.get("course", "")
            departments = record.get("departments", [])
            # If user is not already in users_dict, we create for him empty course list
            if email not in users_dict:
                users_dict[email] = {'all_user_courses': []}


            users_dict[email]['all_user_courses'].append({"course":course,'progress':progress, "status":status})


        all_users_cnt = len(users_dict)
        print_and_log(f"All users amount: {all_users_cnt}", level="INFO")
        
        print_and_log("Processing started!", level="INFO")
        


        for user_item in users_dict.items():
            user, user_data = user_item
            verdict, user = user_analysis(user, user_data)
            

            if verdict == "diligent":
                diligent_users.append(user)
            elif verdict == "undiligent":
                undiligent_users.append(user)
            else:
                mixed_users.append(user)

        print_and_log("Processing finished!", level="INFO")

        diligent_users_cnt = len(diligent_users)
        undiligent_users_cnt = len(undiligent_users)
        mixed_users_cnt = len(mixed_users)

        if diligent_users_cnt + undiligent_users_cnt + mixed_users_cnt != all_users_cnt:
            print_and_log(f"Users count does not match. Dil users {diligent_users_cnt} + Undil users {undiligent_users_cnt} + Mixed users {mixed_users_cnt} != {all_users_cnt}", level="WARNING")

        else:
            print_and_log(f"Users count match. Dil users {diligent_users_cnt} + Undil users {undiligent_users_cnt} + Mixed users {mixed_users_cnt} = {all_users_cnt}", level="INFO")

        print_and_log("Attempting to save data...", level="INFO")
        new_undil, all_undil, fixed_undil = excel(diligent_users, undiligent_users, mixed_users, users_dict, excel_filename, excel_filename_full, bucket_name, region_name)
        timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S (UTC)")
        sns_undiligent(topic_arn, timestamp, new_undil, all_undil, fixed_undil)

    else:
        print_and_log("Status code is not 200!", level="WARNING")

    schedule_name = "bakalauraschedule"
    group_name = 'default'

    next_execution_time(schedule_name, group_name)

    end_time = time.time()
 
    elapsed_time = end_time - start_time
    print_and_log(f'Elapsed time: {elapsed_time:.2f} sec', level="INFO")